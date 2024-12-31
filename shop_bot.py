from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from time import sleep
from platform import system
from random import uniform
import os

main_border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

def set_sheet_columns(active_ws):
    columns = ['Link', 'Título', 'Preço', 'Vendas', 'Fornecedor']
    active_ws.append(columns)
    for row in active_ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=5):
        for cell in row:
            cell.border = main_border_style
            cell.fill = blue_fill

def set_sheet_rows(active_ws, datas):
    for data in datas:
        active_ws.append(data)

    for row in active_ws.iter_rows(min_row=2, max_row=active_ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = main_border_style

browser_options = Options()
browser_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.199 Safari/537.36')
browser_options.add_argument("--disable-blink-features=AutomationControlled")
browser_options.add_argument('--no-sandbox')
browser_options.add_argument('--enable-unsafe-swiftshader')
navigator = webdriver.Chrome(options=browser_options)
aliexpress_search = f'https://pt.aliexpress.com/w/wholesale-PC-Gamer.html?spm=a2g0o.home.search.0'
navigator.get(aliexpress_search)
sleep(uniform(4, 8))
navigator.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
for c in range(0, 4):
    navigator.execute_script("window.scrollTo(0, document.body.scrollHeight)")
    sleep(uniform(2, 4))

WebDriverWait(navigator, 10).until(
    expected_conditions.presence_of_all_elements_located((By.CLASS_NAME, 'search-card-item'))
)

products_datas = navigator.find_elements(By.CLASS_NAME, 'search-card-item')
main_wb = Workbook()
active_ws = main_wb.active
set_sheet_columns(active_ws)
main_rows = []
browser_actions = ActionChains(navigator)
for product in products_datas[:20]:
    browser_actions.move_to_element(product).perform()
    sleep(uniform(4, 8))
    try:
        product_link = product.get_attribute('href')
        product_title = product.find_element(By.CSS_SELECTOR, '.multi--titleText--nXeOvyr').text
        product_price = product.find_element(By.CSS_SELECTOR, '.multi--price-sale--U-S0jtj').text
        sales_elements = product.find_elements(By.CSS_SELECTOR, '.multi--trade--Ktbl2jB')
        product_sales = sales_elements[0].text if sales_elements else 'Não informado'
        product_supplier = product.find_element(By.CSS_SELECTOR, '.cards--storeLink--XkKUQFS').text

    except Exception as e:
        print(f'Erro ao extrair os dados do produto: {product.text}\n\n{e}')

    else:
        main_rows.append([product_link, product_title, product_price, product_sales, product_supplier])

set_sheet_rows(active_ws, main_rows)
for column in active_ws.columns:
    column_letter = column[0].column_letter
    max_width = max(len(str(cell.value)) for cell in column if cell.value)

active_ws.column_dimensions[column_letter].width = max_width + 2
main_wb.save('aliexpress-scrap.xlsx')
sleep(uniform(4, 8))
if system() == 'Windows':
    os.system('start aliexpress-scrap.xlsx')

elif system() == 'Darwin':
    os.system('open aliexpress-scrap.xlsx')

elif system() == 'Linux':
    os.system('xdg-open aliexpress-scrap.xlsx')

else:
    print(f'Sistema operativo não identificado! Acesse o seu arquivo do excel neste local: {os.path.abspath(__file__)}')

navigator.quit()
